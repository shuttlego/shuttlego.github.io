#!/usr/bin/env python3
"""
xlsx → CSV 변환 스크립트.

data/raw/ 내 '삼성전자(수원)'으로 시작하는 xlsx를 파싱하여
data/routes.csv, data/departure_times.csv, data/tmp_stops.csv 를 재생성합니다.
sites.csv 는 수정하지 않습니다.
위경도는 비워 둔 채로 저장합니다. 보강은 scripts/fill_stop_geo.py 를 실행하세요.
"""

from __future__ import annotations

import csv
import re
import sys
from dataclasses import dataclass
from datetime import time as dt_time
from pathlib import Path
from typing import Any, Optional

import openpyxl

# 같은 scripts/ 디렉터리 모듈 (프로젝트 루트에서 실행 시)
if (scripts_dir := Path(__file__).resolve().parent) not in [Path(p) for p in sys.path]:
    sys.path.insert(0, str(scripts_dir))
from fill_stop_geo import fill_stops_from_geo_info, load_geo_info

# ---------------------------------------------------------------------------
# 설정 (프로젝트 루트 기준 data 디렉토리)
# ---------------------------------------------------------------------------

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
RAW_DIR = DATA_DIR / "raw"
SITE_ID = 1
TYPE_출근 = 1
TYPE_퇴근 = 2

# ---------------------------------------------------------------------------
# 데이터 구조
# ---------------------------------------------------------------------------


@dataclass
class Route:
    route_id: int
    site_id: int
    route_name: str
    type_: int  # 1=출근, 2=퇴근
    operator: str
    notes: str


@dataclass
class Stop:
    route_id: int
    sequence: int
    stop_name: str
    lat: str
    lng: str


def normalize_route_name(raw: Any) -> str:
    if raw is None:
        return ""
    s = str(raw).strip().replace("\n", " ").replace("\r", " ")
    return " ".join(s.split())


def normalize_stop_name(s: str) -> str:
    """정류장명에서 큰따옴표 제거. CSV 필드 내부의 \"는 저장 시 이스케이프되며, 원본 xlsx에 \"가 있으면 그대로 들어오므로 제거."""
    if not s:
        return s
    return s.replace('"', '').strip()


def parse_time_cell(value: Any) -> list[str]:
    """xlsx 시간 셀 → 리스트. 단일이면 ['HH:MM'], 범위면 ['HH:MM ~ HH:MM'] 한 개."""
    if value is None:
        return []
    if isinstance(value, dt_time):
        return [value.strftime("%H:%M")]
    s = str(value).strip()
    pattern = r"\b(\d{1,2}):(\d{2})\b"
    matches = re.findall(pattern, s)
    if not matches:
        return []
    if len(matches) == 1:
        h, m = matches[0]
        return [f"{int(h):02d}:{m}"]
    start_h, start_m = int(matches[0][0]), int(matches[0][1])
    end_h, end_m = int(matches[1][0]), int(matches[1][1])
    start_str = f"{start_h:02d}:{start_m:02d}"
    end_str = f"{end_h:02d}:{end_m:02d}"
    end_min = end_h * 60 + end_m
    start_min = start_h * 60 + start_m
    if end_min <= start_min:
        return [start_str]
    return [f"{start_str} ~ {end_str}"]


def parse_stops_출근(cell_value: Any) -> list[str]:
    """출근 탭 정차장 셀 → stop_name 리스트 (순서 유지)."""
    if cell_value is None:
        return []
    text = str(cell_value).strip()
    if not text:
        return []
    names = []
    for line in text.split("\n"):
        line = line.strip()
        if not line:
            continue
        # "1. xxx", "2 yyy"(점 생략), "4 .명지대"(숫자와 점 사이 공백), 전각 마침표(．) 허용
        m = re.match(r"^\d+\s*[.．\s]*\s*(.+)$", line)
        if m:
            names.append(normalize_stop_name(m.group(1)))
        else:
            names.append(normalize_stop_name(line))
    return names


def parse_stops_퇴근(cell_value: Any) -> list[str]:
    """퇴근 탭 정차장 셀 → stop_name 리스트. '(이용 장소 : xx)' 를 1번, 이어서 1. 2. ..."""
    if cell_value is None:
        return []
    text = str(cell_value).strip()
    if not text:
        return []
    names = []
    # "(이용 장소 : 정문 버스승차장)" 또는 "(이용장소: xx)"
    m_place = re.search(r"\(이용\s*장소\s*:\s*([^)]+)\)", text, re.IGNORECASE)
    if m_place:
        names.append(normalize_stop_name(m_place.group(1)))
    lines = text.split("\n")
    for line in lines:
        line = line.strip()
        if not line or re.match(r"^\(이용", line, re.IGNORECASE):
            continue
        m = re.match(r"^\d+\s*[.．\s]*\s*(.+)$", line)
        if m:
            names.append(normalize_stop_name(m.group(1)))
    return names


# ---------------------------------------------------------------------------
# xlsx 파싱
# ---------------------------------------------------------------------------


def collect_routes_from_sheet(ws, sheet_name: str, route_type: int, next_route_id: int):
    """
    한 시트(출근 또는 퇴근)에서 노선/출발시간/정차장 수집.
    next_route_id: 이 시트에서 쓸 시작 route_id.
    반환: (routes, departure_times, stops, next_route_id_after)
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return [], [], [], next_route_id
    # 1행 제목, 2행 헤더, 3행부터 데이터
    data_rows = rows[2:]
    parse_stops = parse_stops_출근 if route_type == TYPE_출근 else parse_stops_퇴근

    routes = []
    departure_times = []  # (route_id, time_str)
    stops = []  # (route_id, sequence, stop_name)

    current_route_id: Optional[int] = None
    current_route_name: Optional[str] = None
    current_notes = ""
    current_route_operators: set[str] = set()  # 노선별 운행업체 수집 (여러 행에 다른 업체 있을 수 있음)

    for row in data_rows:
        if not row or len(row) < 3:
            continue
        col0, col1, col2 = row[0], row[1], row[2]
        operator = (row[3] or "") if len(row) > 3 else ""
        notes = (row[4] or "") if len(row) > 4 else ""
        if isinstance(operator, str):
            operator = operator.strip().replace("\n", " ")
        else:
            operator = str(operator or "").strip()
        if isinstance(notes, str):
            notes = notes.strip().replace("\n", " ")
        else:
            notes = str(notes or "").strip()

        if operator:
            current_route_operators.add(operator)

        if col0 is not None and str(col0).strip():
            # 이전 노선의 operator를 수집한 값으로 반영 (여러 업체면 "/"로 연결)
            if routes:
                routes[-1].operator = "/".join(sorted(current_route_operators))
            current_route_operators = {operator} if operator else set()

            # 새 노선 시작 — 비고는 이 행 값만 사용 (병합 셀은 다른 행에서 None으로 옴)
            current_route_id = next_route_id
            next_route_id += 1
            current_route_name = normalize_route_name(col0)
            current_notes = notes
            routes.append(
                Route(
                    route_id=current_route_id,
                    site_id=SITE_ID,
                    route_name=current_route_name,
                    type_=route_type,
                    operator=operator,  # 나중에 routes[-1].operator 로 덮어씀
                    notes=current_notes,
                )
            )
            # 정차장
            if col1 is not None and str(col1).strip():
                stop_names = parse_stops(col1)
                for seq, name in enumerate(stop_names, start=1):
                    stops.append((current_route_id, seq, name))
                # 출근 노선(type=1)은 최종 정류장에 디지털시티 추가
                if route_type == TYPE_출근:
                    stops.append((current_route_id, len(stop_names) + 1, "디지털시티"))

        if current_route_id is not None and col2 is not None:
            for t in parse_time_cell(col2):
                departure_times.append((current_route_id, t))

    if routes and current_route_operators:
        routes[-1].operator = "/".join(sorted(current_route_operators))

    return routes, departure_times, stops, next_route_id


def load_all_from_xlsx(xlsx_paths: list[Path]):
    """모든 xlsx에서 routes, departure_times, stops 수집."""
    all_routes = []
    all_departure_times = []
    all_stops = []  # (route_id, sequence, stop_name)
    next_rid = 1
    for path in sorted(xlsx_paths):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet_name in ("출근", "퇴근"):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            route_type = TYPE_출근 if sheet_name == "출근" else TYPE_퇴근
            routes, dts, stops, next_rid = collect_routes_from_sheet(
                ws, sheet_name, route_type, next_rid
            )
            all_routes.extend(routes)
            all_departure_times.extend(dts)
            all_stops.extend(stops)
        wb.close()
    return all_routes, all_departure_times, all_stops


# ---------------------------------------------------------------------------
# CSV 출력
# ---------------------------------------------------------------------------


def build_stops(stops_with_names: list[tuple[int, int, str]]) -> list[Stop]:
    """(route_id, sequence, stop_name) 리스트 → Stop 리스트. 위경도는 비워 둔다."""
    return [
        Stop(route_id=rid, sequence=seq, stop_name=name, lat="", lng="")
        for rid, seq, name in stops_with_names
    ]


def write_csvs(routes: list[Route], departure_times: list[tuple[int, str]], stops: list[Stop]):
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    with open(DATA_DIR / "routes.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["route_id", "site_id", "route_name", "type", "operator", "notes"])
        for r in routes:
            w.writerow([r.route_id, r.site_id, r.route_name, r.type_, r.operator or "", r.notes or ""])

    with open(DATA_DIR / "departure_times.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["route_id", "time"])
        for rid, t in departure_times:
            w.writerow([rid, t])

    with open(DATA_DIR / "tmp_stops.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["route_id", "sequence", "stop_name", "lat", "lng"])
        for s in stops:
            w.writerow([s.route_id, s.sequence, s.stop_name, s.lat, s.lng])


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------


def main():
    raw_dir = RAW_DIR
    if not raw_dir.is_dir():
        print(f"디렉토리가 없습니다: {raw_dir}")
        return 1
    xlsx_paths = sorted(raw_dir.glob("삼성전자(수원)*.xlsx"))
    if not xlsx_paths:
        print(f"대상 xlsx가 없습니다: {raw_dir}/삼성전자(수원)*.xlsx")
        return 1

    print(f"xlsx {len(xlsx_paths)}개 로드 중...")
    routes, departure_times, stops_raw = load_all_from_xlsx(xlsx_paths)
    print(f"  노선 {len(routes)}개, 출발시간 {len(departure_times)}개, 정차장 {len(stops_raw)}개")

    stops = build_stops(stops_raw)
    # 1차 위경도 보강: tmp_geo_info.csv 로 채움
    tmp_geo_path = DATA_DIR / "tmp_geo_info.csv"
    geo_info = load_geo_info(tmp_geo_path)
    if geo_info:
        stops_tuples = [(s.route_id, s.sequence, s.stop_name, s.lat, s.lng) for s in stops]
        filled = fill_stops_from_geo_info(stops_tuples, geo_info)
        stops = [Stop(rid, seq, name, lat, lng) for rid, seq, name, lat, lng in filled]
        print(f"  tmp_geo_info.csv 로 1차 위경도 보강 완료")

    write_csvs(routes, departure_times, stops)
    print(f"저장 완료: {DATA_DIR}/routes.csv, departure_times.csv, tmp_stops.csv")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
